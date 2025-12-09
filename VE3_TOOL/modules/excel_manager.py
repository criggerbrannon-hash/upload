"""
Excel Manager Module
====================

This module handles reading and writing prompt data to Excel files.
It manages two sheets:
- "characters": Character definitions and their visual prompts
- "scenes": Scene-by-scene prompts for image and video generation

Usage:
    workbook = PromptWorkbook()
    workbook.load_or_create(Path("project/prompts/KA1-0001_prompts.xlsx"))
    characters = workbook.get_characters()
    workbook.update_scene(scene_id=1, img_prompt="...", status_img="done")
    workbook.save()
"""

import logging
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .utils import get_logger


# ============================================================================
# Data Classes
# ============================================================================

@dataclass
class Character:
    """Represents a character in the story."""
    id: str  # e.g., "nvc", "nvp1"
    role: str  # "main" or "supporting"
    name: str  # Character name in story
    english_prompt: str = ""  # English description for image generation
    vietnamese_prompt: str = ""  # Vietnamese description (optional)
    image_file: str = ""  # Reference image filename
    status: str = "pending"  # pending, done, error


@dataclass
class Scene:
    """Represents a scene in the story."""
    scene_id: int
    srt_start: str = ""  # Start time from SRT
    srt_end: str = ""  # End time from SRT
    srt_text: str = ""  # Subtitle text
    img_prompt: str = ""  # Prompt for image generation
    video_prompt: str = ""  # Prompt for video generation
    img_path: str = ""  # Path to generated image
    video_path: str = ""  # Path to generated video
    status_img: str = "pending"  # pending, generating, done, error
    status_vid: str = "pending"  # pending, generating, done, error


# ============================================================================
# Column Definitions
# ============================================================================

CHARACTER_COLUMNS = [
    "id",
    "role",
    "name",
    "english_prompt",
    "vietnamese_prompt",
    "image_file",
    "status"
]

SCENE_COLUMNS = [
    "scene_id",
    "srt_start",
    "srt_end",
    "srt_text",
    "img_prompt",
    "video_prompt",
    "img_path",
    "video_path",
    "status_img",
    "status_vid"
]


# ============================================================================
# Excel Styling
# ============================================================================

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Column widths
CHARACTER_WIDTHS = {
    "id": 10,
    "role": 12,
    "name": 20,
    "english_prompt": 60,
    "vietnamese_prompt": 40,
    "image_file": 20,
    "status": 12
}

SCENE_WIDTHS = {
    "scene_id": 10,
    "srt_start": 15,
    "srt_end": 15,
    "srt_text": 50,
    "img_prompt": 60,
    "video_prompt": 60,
    "img_path": 30,
    "video_path": 30,
    "status_img": 12,
    "status_vid": 12
}


# ============================================================================
# PromptWorkbook Class
# ============================================================================

class PromptWorkbook:
    """
    Manages the Excel workbook for storing prompts and tracking generation status.

    Attributes:
        path: Path to the Excel file.
        workbook: openpyxl Workbook instance.
        logger: Logger instance.
    """

    def __init__(self, logger: Optional[logging.Logger] = None):
        """
        Initialize PromptWorkbook.

        Args:
            logger: Optional logger instance.
        """
        self.path: Optional[Path] = None
        self.workbook: Optional[Workbook] = None
        self.logger = logger or get_logger("ve3_tool.excel_manager")

    def load_or_create(self, path: Path) -> 'PromptWorkbook':
        """
        Load existing workbook or create a new one.

        Args:
            path: Path to the Excel file.

        Returns:
            Self for method chaining.
        """
        self.path = Path(path)

        # Ensure directory exists
        self.path.parent.mkdir(parents=True, exist_ok=True)

        if self.path.exists():
            self.logger.info(f"Loading existing workbook: {self.path.name}")
            self.workbook = load_workbook(self.path)
            self._ensure_sheets_exist()
        else:
            self.logger.info(f"Creating new workbook: {self.path.name}")
            self.workbook = Workbook()
            self._create_sheets()
            self.save()

        return self

    def _create_sheets(self) -> None:
        """Create the required sheets with headers."""
        # Remove default sheet if it exists
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]

        # Create characters sheet
        char_sheet = self.workbook.create_sheet("characters", 0)
        self._setup_sheet_headers(char_sheet, CHARACTER_COLUMNS, CHARACTER_WIDTHS)

        # Create scenes sheet
        scene_sheet = self.workbook.create_sheet("scenes", 1)
        self._setup_sheet_headers(scene_sheet, SCENE_COLUMNS, SCENE_WIDTHS)

    def _setup_sheet_headers(
        self,
        sheet: Worksheet,
        columns: list[str],
        widths: dict[str, int]
    ) -> None:
        """Set up headers for a sheet."""
        for col_idx, col_name in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

            # Set column width
            col_letter = cell.column_letter
            sheet.column_dimensions[col_letter].width = widths.get(col_name, 15)

        # Freeze header row
        sheet.freeze_panes = "A2"

    def _ensure_sheets_exist(self) -> None:
        """Ensure required sheets exist in loaded workbook."""
        if "characters" not in self.workbook.sheetnames:
            char_sheet = self.workbook.create_sheet("characters", 0)
            self._setup_sheet_headers(char_sheet, CHARACTER_COLUMNS, CHARACTER_WIDTHS)

        if "scenes" not in self.workbook.sheetnames:
            scene_sheet = self.workbook.create_sheet("scenes", 1)
            self._setup_sheet_headers(scene_sheet, SCENE_COLUMNS, SCENE_WIDTHS)

    def save(self) -> None:
        """Save the workbook to disk."""
        if self.workbook and self.path:
            self.workbook.save(self.path)
            self.logger.debug(f"Workbook saved: {self.path.name}")

    # ========================================================================
    # Character Methods
    # ========================================================================

    def get_characters(self) -> list[Character]:
        """
        Get all characters from the workbook.

        Returns:
            List of Character objects.
        """
        sheet = self.workbook["characters"]
        characters = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue

            characters.append(Character(
                id=str(row[0] or ""),
                role=str(row[1] or ""),
                name=str(row[2] or ""),
                english_prompt=str(row[3] or ""),
                vietnamese_prompt=str(row[4] or ""),
                image_file=str(row[5] or ""),
                status=str(row[6] or "pending")
            ))

        return characters

    def add_character(self, character: Character) -> None:
        """
        Add a new character to the workbook.

        Args:
            character: Character object to add.
        """
        sheet = self.workbook["characters"]
        next_row = sheet.max_row + 1

        values = [
            character.id,
            character.role,
            character.name,
            character.english_prompt,
            character.vietnamese_prompt,
            character.image_file,
            character.status
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=next_row, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

        self.logger.debug(f"Added character: {character.id}")

    def update_character(
        self,
        char_id: str,
        **kwargs: Any
    ) -> bool:
        """
        Update a character by ID.

        Args:
            char_id: Character ID to update.
            **kwargs: Fields to update (english_prompt, status, etc.)

        Returns:
            True if character was found and updated, False otherwise.
        """
        sheet = self.workbook["characters"]
        col_indices = {col: idx + 1 for idx, col in enumerate(CHARACTER_COLUMNS)}

        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == char_id:
                for field_name, value in kwargs.items():
                    if field_name in col_indices:
                        sheet.cell(
                            row=row_idx,
                            column=col_indices[field_name],
                            value=value
                        )
                self.logger.debug(f"Updated character: {char_id}")
                return True

        return False

    def clear_characters(self) -> None:
        """Remove all characters (keep headers)."""
        sheet = self.workbook["characters"]
        for row_idx in range(sheet.max_row, 1, -1):
            sheet.delete_rows(row_idx)

    # ========================================================================
    # Scene Methods
    # ========================================================================

    def get_scenes(self) -> list[Scene]:
        """
        Get all scenes from the workbook.

        Returns:
            List of Scene objects.
        """
        sheet = self.workbook["scenes"]
        scenes = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue

            scenes.append(Scene(
                scene_id=int(row[0]) if row[0] else 0,
                srt_start=str(row[1] or ""),
                srt_end=str(row[2] or ""),
                srt_text=str(row[3] or ""),
                img_prompt=str(row[4] or ""),
                video_prompt=str(row[5] or ""),
                img_path=str(row[6] or ""),
                video_path=str(row[7] or ""),
                status_img=str(row[8] or "pending"),
                status_vid=str(row[9] or "pending")
            ))

        return scenes

    def add_scene(self, scene: Scene) -> None:
        """
        Add a new scene to the workbook.

        Args:
            scene: Scene object to add.
        """
        sheet = self.workbook["scenes"]
        next_row = sheet.max_row + 1

        values = [
            scene.scene_id,
            scene.srt_start,
            scene.srt_end,
            scene.srt_text,
            scene.img_prompt,
            scene.video_prompt,
            scene.img_path,
            scene.video_path,
            scene.status_img,
            scene.status_vid
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=next_row, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

        self.logger.debug(f"Added scene: {scene.scene_id}")

    def update_scene(
        self,
        scene_id: int,
        **kwargs: Any
    ) -> bool:
        """
        Update a scene by ID.

        Args:
            scene_id: Scene ID to update.
            **kwargs: Fields to update (img_prompt, status_img, etc.)

        Returns:
            True if scene was found and updated, False otherwise.
        """
        sheet = self.workbook["scenes"]
        col_indices = {col: idx + 1 for idx, col in enumerate(SCENE_COLUMNS)}

        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value is not None and int(cell_value) == scene_id:
                for field_name, value in kwargs.items():
                    if field_name in col_indices:
                        sheet.cell(
                            row=row_idx,
                            column=col_indices[field_name],
                            value=value
                        )
                self.logger.debug(f"Updated scene: {scene_id}")
                return True

        return False

    def clear_scenes(self) -> None:
        """Remove all scenes (keep headers)."""
        sheet = self.workbook["scenes"]
        for row_idx in range(sheet.max_row, 1, -1):
            sheet.delete_rows(row_idx)

    def get_scenes_by_status(
        self,
        status_img: Optional[str] = None,
        status_vid: Optional[str] = None
    ) -> list[Scene]:
        """
        Get scenes filtered by status.

        Args:
            status_img: Filter by image status (pending, done, error).
            status_vid: Filter by video status.

        Returns:
            List of Scene objects matching the filters.
        """
        scenes = self.get_scenes()

        if status_img is not None:
            scenes = [s for s in scenes if s.status_img == status_img]

        if status_vid is not None:
            scenes = [s for s in scenes if s.status_vid == status_vid]

        return scenes

    # ========================================================================
    # Utility Methods
    # ========================================================================

    def has_prompts(self) -> bool:
        """
        Check if the workbook has any prompts filled in.

        Returns:
            True if at least one scene has an img_prompt.
        """
        scenes = self.get_scenes()
        return any(s.img_prompt for s in scenes)

    def get_statistics(self) -> dict[str, Any]:
        """
        Get statistics about the workbook content.

        Returns:
            Dictionary with statistics.
        """
        scenes = self.get_scenes()
        characters = self.get_characters()

        return {
            "total_characters": len(characters),
            "total_scenes": len(scenes),
            "scenes_with_img_prompt": sum(1 for s in scenes if s.img_prompt),
            "scenes_with_video_prompt": sum(1 for s in scenes if s.video_prompt),
            "images_done": sum(1 for s in scenes if s.status_img == "done"),
            "images_pending": sum(1 for s in scenes if s.status_img == "pending"),
            "images_error": sum(1 for s in scenes if s.status_img == "error"),
            "videos_done": sum(1 for s in scenes if s.status_vid == "done"),
            "videos_pending": sum(1 for s in scenes if s.status_vid == "pending"),
            "videos_error": sum(1 for s in scenes if s.status_vid == "error"),
        }

    def export_prompts_text(self, output_path: Path) -> None:
        """
        Export all prompts to a plain text file for review.

        Args:
            output_path: Path to save the text file.
        """
        scenes = self.get_scenes()
        characters = self.get_characters()

        lines = []
        lines.append("=" * 80)
        lines.append("CHARACTERS")
        lines.append("=" * 80)

        for char in characters:
            lines.append(f"\n[{char.id}] {char.name} ({char.role})")
            lines.append(f"Prompt: {char.english_prompt}")
            lines.append(f"Image: {char.image_file}")
            lines.append("-" * 40)

        lines.append("\n" + "=" * 80)
        lines.append("SCENES")
        lines.append("=" * 80)

        for scene in scenes:
            lines.append(f"\n[Scene {scene.scene_id}] {scene.srt_start} - {scene.srt_end}")
            lines.append(f"Text: {scene.srt_text[:100]}...")
            lines.append(f"Image Prompt: {scene.img_prompt}")
            lines.append(f"Video Prompt: {scene.video_prompt}")
            lines.append(f"Status: img={scene.status_img}, vid={scene.status_vid}")
            lines.append("-" * 40)

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        self.logger.info(f"Exported prompts to: {output_path}")
